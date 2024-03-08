import openpyxl
import xlrd
import os
import difflib


# Mapping of variations to application options
applications_mapping = {
    'rnaseq': 'RNAseq',
    'trnaseq': 'tRNAseq',
    'mrnaseq': 'mRNAseq', 
    '3mrnaseq': '3mRNAseq', 
    'chipseq': 'ChIPseq', 
    'atacseq': 'ATACseq', 
    'ampliseq': 'ampliseq', 
    'scrnaseq': 'scRNAseq',    
    'scvdjseq': 'scVDJseq', 
    'scatacseq': 'scATACseq', 
    'mirnaseq': 'miRNAseq', 
    'bwgs': 'BWGS', 
    'wes': 'WES', 
    'fastq': 'fastq', 
    '16s': '16S', 
    'mag': 'MAG'
    }

def parse_sciebo_report(report_path):
    if report_path.lower().endswith(".xlsx"):
        return parse_sciebo_xlsx_report(report_path)
    elif report_path.lower().endswith(".xls"):
        return parse_sciebo_xls_report(report_path)
    else:
        return BaseException("unknown file format for sciebo_report")
    
def parse_sciebo_xls_report(report_path):
    """ Gather:
    - Project Name
    - Sequencing Kit
    - Cycles Read 1/2
    - Cycles Index 1/2
    - Density
    - Clusters PF
    - Yield
    - % >= Q30
    """
    sequencing_kit = None
    cycles_read_1 = None
    cycles_index_1 = None
    cycles_read_2 = None
    cycles_index_2 = None
    density = None
    clusters_pf = None
    yields = None
    q_30 = None
    project_name = None
    protocol_name = None
    application = None
    phix_input = None

    file_name = os.path.splitext(os.path.basename(report_path))[0]
    parts = file_name.split('_')
    # Extract the required parts
    date_part = parts[0]  # YYMMDD
    # Convert the date parts to integers for comparison
    year = int(date_part[:2]) + 2000
    month = int(date_part[2:4])
    day = int(date_part[4:6])

    application_part = parts[-1]  # Application
    # Combine the required parts to form the desired string
    protocol_name = f"{date_part}_{application_part}"

    # application = applications_mapping.get(application_part.lower(), 'unknown')
    # Find the closest match using difflib
    closest_match = difflib.get_close_matches(application_part.lower(), applications_mapping.keys(), n=1, cutoff=0.6)
    # If a close match is found, use the corresponding option; otherwise, use 'unknown'
    application = applications_mapping.get(closest_match[0], 'unknown') if closest_match else 'unknown'


    # Open the Excel file
    workbook = xlrd.open_workbook(report_path)

    # Iterate through sheets
    for sheet_name in workbook.sheet_names():
        sheet = workbook.sheet_by_name(sheet_name)

        # Iterate through rows and columns
        for i in range(sheet.nrows):
            for j in range(sheet.ncols):
                cell_value = sheet.cell_value(i, j)

                if cell_value is None:
                    continue
                if cell_value == "Cycles Read 1":
                    cycles_read_1 = sheet.cell_value(i, j + 1)
                elif cell_value == "Cycles Index 1":
                    cycles_index_1 = sheet.cell_value(i, j + 1)
                elif cell_value == "Cycles Index 2":
                    cycles_index_2 = sheet.cell_value(i, j + 1)
                elif cell_value == "Cycles Read 2":
                    cycles_read_2 = sheet.cell_value(i, j + 1)
                elif cell_value == "Density":
                    density = sheet.cell_value(i, j + 1)
                elif cell_value == "Clusters PF":
                    clusters_pf = sheet.cell_value(i, j + 1)
                elif cell_value == "Yield":
                    yields = sheet.cell_value(i, j + 1)
                    if type(yields) is str:
                        yields = yields.replace(',', '.')
                        yields =''.join(char for char in yields if char.isdigit() or char == '.')
                elif cell_value == "% >= Q30":
                    q_30 = sheet.cell_value(i, j + 1)
                    if q_30 is None:
                        continue
                    try:
                        if type(q_30) is str:
                            q_30 = q_30.replace('%', '')
                            q_30 = float(q_30.replace(' ',''))
                        if q_30 < 1:
                            q_30 *= 100
                    except BaseException:
                        print(f"q_30 = {q_30} was not handled successfully")
                elif isinstance(cell_value, str) and "kit" in cell_value.lower():
                    # Find the position of ':' and get the substring after it
                    index_colon = cell_value.find(':')
                    if index_colon != -1:
                        sequencing_kit = cell_value[index_colon + 1:].strip()
                elif isinstance(cell_value, str) and "project" in cell_value.lower():
                    # Find the position of ':' and get the substring after it
                    index_colon = cell_value.find(':')
                    if index_colon != -1:
                        project_name = cell_value[index_colon + 1:].strip()
                elif isinstance(cell_value, str) and "phix" in cell_value.lower():
                    if (year, month, day) > (2024, 1, 16):
                        # The date is past 16.01.24 and the phix protocol is in place
                        phix_input = sheet.cell_value(i, j + 1)
                    else:
                        # No Protocol prior to the 16.01.24
                        phix_input = None

    return [sequencing_kit, cycles_read_1, cycles_index_1, cycles_read_2, cycles_index_2, density, clusters_pf, yields, q_30 , project_name, protocol_name, application, phix_input]
    

def parse_sciebo_xlsx_report(report_path):
    """ Gather:
    - Project Name
    - Sequencing Kit
    - Cycles Read 1/2
    - Cycles Index 1/2
    - Density
    - Clusters PF
    - Yield
    - % >= Q30
    """
    sequencing_kit = None
    cycles_read_1 = None
    cycles_index_1 = None
    cycles_read_2 = None
    cycles_index_2 = None
    density = None
    clusters_pf = None
    yields = None
    q_30 = None
    project_name = None
    protocol_name = None
    application = None
    phix_input = None

    file_name = os.path.splitext(os.path.basename(report_path))[0]
    parts = file_name.split('_')
    # Extract the required parts
    date_part = parts[0]  # YYMMDD
    # Convert the date parts to integers for comparison
    year = int(date_part[:2]) + 2000
    month = int(date_part[2:4])
    day = int(date_part[4:6])

    application_part = parts[-1]  # Application
    # Combine the required parts to form the desired string
    protocol_name = f"{date_part}_{application_part}"
    
    # application = applications_mapping.get(application_part.lower(), 'unknown')
    # Find the closest match using difflib
    closest_match = difflib.get_close_matches(application_part.lower(), applications_mapping.keys(), n=1, cutoff=0.6)
    # If a close match is found, use the corresponding option; otherwise, use 'unknown'
    application = applications_mapping.get(closest_match[0], 'unknown') if closest_match else 'unknown'

    wb = openpyxl.load_workbook(filename=report_path, data_only=True)

    # Get All Sheets
    for excel_sheet_name in wb.sheetnames:

        # Get Sheet Object by names
        excel_sheet = wb[excel_sheet_name]

        for i in range(1,excel_sheet.max_row):
            for j in range(1, excel_sheet.max_column):
                excel_sheet_name_cell = excel_sheet.cell(row=i, column=j).value
                if excel_sheet_name_cell == None:
                    continue
                elif excel_sheet_name_cell == "Cycles Read 1":
                    cycles_read_1 = excel_sheet.cell(row=i, column=j+1).value
                elif excel_sheet_name_cell == "Cycles Index 1":
                    cycles_index_1 = excel_sheet.cell(row=i, column=j+1).value
                elif excel_sheet_name_cell == "Cycles Index 2":
                    cycles_index_2 = excel_sheet.cell(row=i, column=j+1).value
                elif excel_sheet_name_cell == "Cycles Read 2":
                    cycles_read_2 = excel_sheet.cell(row=i, column=j+1).value
                elif excel_sheet_name_cell == "Density":
                    density = excel_sheet.cell(row=i, column=j+1).value
                elif excel_sheet_name_cell == "Clusters PF":
                    clusters_pf = excel_sheet.cell(row=i, column=j+1).value
                elif excel_sheet_name_cell == "Yield":
                    yields = excel_sheet.cell(row=i, column=j+1).value
                    if type(yields) is str:
                        yields = yields.replace(',', '.')
                        yields =''.join(char for char in yields if char.isdigit() or char == '.')
                elif excel_sheet_name_cell == "% >= Q30":
                    q_30 = excel_sheet.cell(row=i, column=j+1).value
                    if q_30 is None:
                        continue
                    try:
                        if type(q_30) is str:
                            q_30 = q_30.replace('%', '')
                            q_30 = float(q_30.replace(' ',''))
                        if q_30 < 1:
                            q_30 *= 100
                        q_30 = str(q_30)
                    except BaseException:
                        print(f"q_30 = {q_30} was not handled successfully")
                elif type(excel_sheet_name_cell) == str and "phix" in excel_sheet_name_cell.lower():
                    if (year, month, day) > (2024, 1, 16):
                        # The date is past 16.01.24 and the phix protocol is in place
                        phix_input = excel_sheet.cell(row=i, column=j+1).value
                    else:
                        # No Protocol prior to the 16.01.24
                        phix_input = None
                elif type(excel_sheet_name_cell) == str and "kit" in excel_sheet_name_cell.lower():
                    # Find the position of ':' and get the substring after it
                    index_colon = excel_sheet_name_cell.find(':')
                    if index_colon != -1:
                        sequencing_kit = excel_sheet_name_cell[index_colon + 1:].strip()
                elif type(excel_sheet_name_cell) == str and "project" in excel_sheet_name_cell.lower():
                    # Find the position of ':' and get the substring after it
                    index_colon = excel_sheet_name_cell.find(':')
                    if index_colon != -1:
                        project_name = excel_sheet_name_cell[index_colon + 1:].strip()
                    
    return [sequencing_kit, cycles_read_1, cycles_index_1, cycles_read_2, cycles_index_2, density, clusters_pf, yields, q_30 , project_name, protocol_name, application, phix_input]

